# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "requests>=2.31"]
# ///
"""Push not-yet-sent rows from an .xlsx to an API, then mark them sent.

Author: Catalin (shared via the Arona intelligence center).

Idempotent: a row is sent only if its status cell is empty; on success the
status cell is stamped with an ISO timestamp, so re-running skips delivered rows.

Robustness:
- Reads cell *values* from a data_only handle, so formula cells send their
  computed result (not the formula text). Requires the file to have been saved
  by Excel/LibreOffice (cached values present).
- Writes the "sent" marker into a separate writable handle so the workbook's
  formulas are preserved on save.
- Coerces date/datetime/time cells to ISO strings before POSTing.
"""
import argparse
import datetime
import json
import sys

import openpyxl
import requests


def _jsonable(value):
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def main() -> int:
    ap = argparse.ArgumentParser(description="Push xlsx rows to an API and mark them sent.")
    ap.add_argument("xlsx", help="path to the .xlsx file (e.g. $NAS_ROOT/data/orders.xlsx)")
    ap.add_argument("api_url", help="target API endpoint (receives one JSON object per row)")
    ap.add_argument("--sheet", default=None, help="worksheet name (default: first sheet)")
    ap.add_argument("--status-col", default="sent_at", help="header of the 'sent' marker column")
    ap.add_argument("--timeout", type=float, default=30.0)
    ap.add_argument("--dry-run", action="store_true", help="print rows, change nothing")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb[args.sheet] if args.sheet else wb.active
    wb_vals = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws_vals = wb_vals[args.sheet] if args.sheet else wb_vals.active

    headers = [c.value for c in ws[1]]
    if args.status_col not in headers:
        headers.append(args.status_col)
        ws.cell(row=1, column=len(headers), value=args.status_col)
    status_idx = headers.index(args.status_col) + 1

    sent = skipped = failed = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=status_idx).value:
            skipped += 1
            continue

        payload = {
            headers[c]: _jsonable(ws_vals.cell(row=r, column=c + 1).value)
            for c in range(len(headers))
            if headers[c] and headers[c] != args.status_col
        }
        if all(v is None for v in payload.values()):
            skipped += 1
            continue

        if args.dry_run:
            print("DRY", json.dumps(payload, default=str))
            sent += 1
            continue

        try:
            resp = requests.post(args.api_url, json=payload, timeout=args.timeout)
            resp.raise_for_status()
        except Exception as exc:
            sys.stderr.write(f"row {r}: {exc}\n")
            failed += 1
            continue

        ws.cell(row=r, column=status_idx,
                value=datetime.datetime.now().isoformat(timespec="seconds"))
        sent += 1

    if not args.dry_run:
        wb.save(args.xlsx)

    sys.stderr.write(f"sent={sent} skipped={skipped} failed={failed}\n")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
