# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "psycopg2-binary", "requests>=2.31"]
# ///
"""
siruta_sync.py — îmbogățește nomenclatorul RO cu SIRUTA oficial (INS, data.gov.ro).

Ce face: descarcă registrul oficial al localităților RO (SIRUTA 2025, ~17k rânduri: județ →
UAT municipiu/oraș/comună → localitate/sat, cu cod SIRUTA + cod poștal), îl încarcă în tabelul
metrics `romania_siruta` (registrul COMPLET, autoritativ) și leagă `romania_addresses.cod_siruta`
pe (județ, localitate). Astfel validatorul RO (`address_nomenclator.py`) nu mai supra-respinge
localități reale care lipsesc din tabelul postal parțial (romania_addresses ~11k localități vs
SIRUTA ~13.7k localități reale).

ADITIV + IDEMPOTENT: CREATE TABLE IF NOT EXISTS · ADD COLUMN IF NOT EXISTS · UPSERT ON CONFLICT ·
UPDATE (doar setează cod_siruta). NICIUN DELETE/TRUNCATE/DROP. Se poate rerula oricând.

  uv run siruta_sync.py               # DRY-RUN: descarcă, parsează, arată câștigul de acoperire (nu scrie)
  uv run siruta_sync.py --apply       # scrie în metrics (romania_siruta + romania_addresses.cod_siruta)
  uv run siruta_sync.py --year 2025   # altă ediție SIRUTA de pe data.gov.ro
Read-only fără --apply. Scrie DOAR în metrics (romania_siruta, romania_addresses).
"""
import os, re, sys, io, argparse, subprocess, unicodedata, tempfile
import urllib.parse as up
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = os.path.dirname(os.path.abspath(__file__))
KB = os.path.join(HERE, "..", "..", "..", "core", "scripts", "kb.py")
CKAN = "https://data.gov.ro/api/3/action/package_show?id=siruta-{year}"


# ===== normalizare (IDENTICĂ cu address_nomenclator.py — trebuie să se lege pe judet_norm/localitate_norm) =====
def strip_diacritics(s):
    if not s: return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return (s.replace("ș", "s").replace("ş", "s").replace("ț", "t").replace("ţ", "t")
             .replace("ă", "a").replace("â", "a").replace("î", "i"))
def norm_text(s):
    s = strip_diacritics(s or "").lower()
    s = re.sub(r"[',’`\"“”]", " ", s)
    s = re.sub(r"[,.;:()_/\\\-]+", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

# prefixe administrative de scos ca să obții numele localității pe care-l tastează clientul
ADMIN_PREFIX = re.compile(r"^(judetul|municipiul|orasul|oras|comuna|satul|sat|sectorul|sector)\s+")
def loc_norm(denloc):
    return ADMIN_PREFIX.sub("", norm_text(denloc)).strip() or norm_text(denloc)


def secret(k):
    return os.environ.get(k) or subprocess.run(["uv", "run", KB, "secret-get", k],
                                               capture_output=True, text=True).stdout.strip()


def download_siruta(year):
    r = requests.get(CKAN.format(year=year), timeout=60); r.raise_for_status()
    res = r.json()["result"]["resources"]
    # fișierul e XLSX chiar dacă e etichetat CSV pe portal → caută after ext/nume
    cand = [x for x in res if "siruta" in (x.get("name") or "").lower() and (x.get("url") or "").lower().endswith((".csv", ".xlsx"))]
    if not cand:
        cand = [x for x in res if (x.get("format") or "").upper() in ("CSV", "XLSX")]
    url = cand[0]["url"]
    print("  descarc:", url.split("/")[-1])
    data = requests.get(url, timeout=120); data.raise_for_status()
    fp = os.path.join(tempfile.gettempdir(), "siruta_%s.xlsx" % year)
    open(fp, "wb").write(data.content)
    return fp


def parse_siruta(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip().upper() for h in next(it)]
    col = {h: i for i, h in enumerate(hdr)}
    need = ["SIRUTA", "DENLOC", "CODP", "JUD", "SIRSUP", "TIP", "NIV", "MED"]
    for n in need:
        if n not in col: raise SystemExit("SIRUTA: lipsește coloana %s (am %s)" % (n, hdr))
    rows = []
    jud_name = {}   # cod județ -> nume (din rândurile NIV=1)
    for r in it:
        if not r or r[col["SIRUTA"]] in (None, ""): continue
        cod = int(r[col["SIRUTA"]]); niv = int(r[col["NIV"]] or 0); jud = int(r[col["JUD"]] or 0)
        den = str(r[col["DENLOC"]] or "").strip()
        codp = str(r[col["CODP"]] or "").strip()
        codp = codp if (codp and codp != "0") else None
        rows.append({"cod": cod, "den": den, "codp": codp, "jud": jud,
                     "sirsup": int(r[col["SIRSUP"]] or 0), "tip": int(r[col["TIP"]] or 0),
                     "niv": niv, "med": str(r[col["MED"]] or "").strip(),
                     "nuts": str(r[col["NUTS"]] or "").strip() if "NUTS" in col else None})
        if niv == 1:  # rând-antet județ: "JUDEŢUL ALBA" / "MUNICIPIUL BUCUREŞTI"
            jud_name[jud] = ADMIN_PREFIX.sub("", norm_text(den)).strip() or norm_text(den)
    for x in rows:
        x["judet_norm"] = jud_name.get(x["jud"], "")
        x["denumire_norm"] = norm_text(x["den"])
        x["localitate_norm"] = loc_norm(x["den"])
    return rows, jud_name


DDL = """
CREATE TABLE IF NOT EXISTS public.romania_siruta (
  cod_siruta      bigint PRIMARY KEY,
  denumire        text,
  denumire_norm   text,
  localitate_norm text,
  tip             int,
  niv             int,
  med             text,
  cod_postal      text,
  sirsup          bigint,
  jud             int,
  judet_norm      text,
  nuts            text
);
CREATE INDEX IF NOT EXISTS idx_siruta_jud_loc ON public.romania_siruta (judet_norm, localitate_norm);
CREATE INDEX IF NOT EXISTS idx_siruta_loc     ON public.romania_siruta (localitate_norm);
CREATE INDEX IF NOT EXISTS idx_siruta_niv     ON public.romania_siruta (niv);
"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="scrie în metrics (altfel doar dry-run)")
    ap.add_argument("--year", default="2025")
    a = ap.parse_args()
    import psycopg2, psycopg2.extras

    print("═" * 70); print("  SIRUTA sync — îmbogățire nomenclator RO (%s)" % ("APPLY" if a.apply else "DRY-RUN")); print("═" * 70)
    rows, jud_name = parse_siruta(download_siruta(a.year))
    by_niv = {}
    for x in rows: by_niv[x["niv"]] = by_niv.get(x["niv"], 0) + 1
    print("  parsat: %d rânduri | județe(NIV1)=%d · UAT(NIV2)=%d · localități(NIV3)=%d" % (
        len(rows), by_niv.get(1, 0), by_niv.get(2, 0), by_niv.get(3, 0)))

    dsn = secret("DATABASE_URL_METRICS"); p = up.urlsplit(dsn)
    cn = psycopg2.connect(up.urlunsplit((p.scheme, p.netloc, p.path, "", ""))); cur = cn.cursor()

    # câștig de acoperire: localități SIRUTA (NIV 2/3) care NU-s în romania_addresses
    cur.execute("SELECT DISTINCT judet_norm, localitate_norm FROM public.romania_addresses")
    have = set((jn, ln) for jn, ln in cur.fetchall())
    siruta_locs = set((x["judet_norm"], x["localitate_norm"]) for x in rows if x["niv"] in (2, 3) and x["localitate_norm"])
    new_locs = siruta_locs - have
    print("  romania_addresses acoperă %d perechi (județ,localitate)" % len(have))
    print("  SIRUTA are %d localități (NIV 2/3) → %d NOI (lipsesc din tabelul actual) = câștig de acoperire" % (len(siruta_locs), len(new_locs)))
    ex = list(new_locs)[:8]
    if ex: print("  ex. localități recuperate:", ", ".join("%s/%s" % (jn, ln) for jn, ln in ex))

    if not a.apply:
        cn.close()
        print("\n  DRY-RUN — nu am scris nimic. Rulează cu --apply ca să încarci romania_siruta + să legi cod_siruta.")
        return

    cur.execute(DDL)
    data = [(x["cod"], x["den"], x["denumire_norm"], x["localitate_norm"], x["tip"], x["niv"],
             x["med"], x["codp"], x["sirsup"], x["jud"], x["judet_norm"], x["nuts"]) for x in rows]
    psycopg2.extras.execute_values(cur,
        """INSERT INTO public.romania_siruta
           (cod_siruta,denumire,denumire_norm,localitate_norm,tip,niv,med,cod_postal,sirsup,jud,judet_norm,nuts)
           VALUES %s ON CONFLICT (cod_siruta) DO UPDATE SET
             denumire=EXCLUDED.denumire, denumire_norm=EXCLUDED.denumire_norm,
             localitate_norm=EXCLUDED.localitate_norm, tip=EXCLUDED.tip, niv=EXCLUDED.niv,
             med=EXCLUDED.med, cod_postal=EXCLUDED.cod_postal, sirsup=EXCLUDED.sirsup,
             jud=EXCLUDED.jud, judet_norm=EXCLUDED.judet_norm, nuts=EXCLUDED.nuts""",
        data, page_size=2000)
    print("  ✓ romania_siruta: %d rânduri upsert" % len(data))

    # leagă cod_siruta pe romania_addresses (aditiv — doar setează, preferă localitatea NIV=3)
    cur.execute("ALTER TABLE public.romania_addresses ADD COLUMN IF NOT EXISTS cod_siruta bigint")
    cur.execute("""
        UPDATE public.romania_addresses ra SET cod_siruta = s.cod_siruta
        FROM (SELECT DISTINCT ON (judet_norm, localitate_norm) judet_norm, localitate_norm, cod_siruta
              FROM public.romania_siruta WHERE niv IN (3,2) AND localitate_norm<>''
              ORDER BY judet_norm, localitate_norm, niv DESC) s
        WHERE ra.judet_norm = s.judet_norm AND ra.localitate_norm = s.localitate_norm""")
    linked = cur.rowcount
    cn.commit()
    cur.execute("SELECT count(*) FROM public.romania_addresses WHERE cod_siruta IS NOT NULL")
    tot_linked = cur.fetchone()[0]
    print("  ✓ romania_addresses.cod_siruta: %d rânduri legate acum (%d în total au cod SIRUTA)" % (linked, tot_linked))
    cn.close()
    print("\n  GATA. Validatorul poate consulta acum romania_siruta pt existența localității (nu mai supra-respinge).")


if __name__ == "__main__":
    main()
