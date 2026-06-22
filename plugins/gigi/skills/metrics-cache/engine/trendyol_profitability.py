#!/usr/bin/env python3
"""
Trendyol Profitability Calculator
──────────────────────────────────
Pulls financial data from Trendyol API (settlements, orders, products)
and calculates profitability per product and per period.

Usage:
    python trendyol_profitability.py [--days 30] [--cost-file costs.csv]

The script outputs:
  - Per-product profitability breakdown
  - Summary statistics (total revenue, commissions, net profit)
  - Excel export with detailed data
"""

import argparse
import csv
import json
import os
import sys
import time
from datetime import datetime, timedelta
from io import StringIO

import requests
from dotenv import load_dotenv

load_dotenv()

# ─── Trendyol API Config (din .env — fără secrete hardcodate) ───────
SELLER_ID = os.environ.get("TRENDYOL_SELLER_ID", "")
API_KEY = os.environ.get("TRENDYOL_API_KEY", "")
API_SECRET = os.environ.get("TRENDYOL_API_SECRET", "")
TOKEN = os.environ.get("TRENDYOL_TOKEN", "")

# International sellers (Romania) need storeFrontCode header
STORE_FRONT_CODE = os.environ.get("TRENDYOL_STORE_FRONT_CODE", "RO")

BASE_URL = "https://apigw.trendyol.com"
INTEGRATION_BASE = f"{BASE_URL}/integration"
SAPIGW_BASE = f"{BASE_URL}/sapigw"

HEADERS = {
    "Authorization": f"Basic {TOKEN}",
    "User-Agent": f"{SELLER_ID} - SelfIntegration",
    "Content-Type": "application/json",
    "storeFrontCode": STORE_FRONT_CODE,
}

# Max 15 days per finance request
MAX_FINANCE_WINDOW_DAYS = 14
# Orders API gives incorrect results with > 7 day windows
MAX_ORDER_WINDOW_DAYS = 7
# Max 50 req / 10 sec — use 0.5s to stay well under limit
RATE_LIMIT_DELAY = 0.5
# Max retries on 429 rate limit before giving up
MAX_429_RETRIES = 10
# Finance API only accepts size=500
FINANCE_PAGE_SIZE = 500
# TVA Romania = 21% (din profit_core, single-source)  |  Transport TVA = 0%
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
try:
    import profit_core as _pc
    TVA_RATE = _pc.vat_for_country("RO")
except Exception:
    TVA_RATE = 0.21
TVA_TRANSPORT = 0.0


# ─── API Helpers ─────────────────────────────────────────────────────
session = requests.Session()
session.headers.update(HEADERS)


def api_get(url, params=None, label="", _retry_count=0):
    """Make a GET request with rate limiting and error handling."""
    time.sleep(RATE_LIMIT_DELAY)
    try:
        r = session.get(url, params=params, timeout=30)
        if r.status_code == 200:
            return r.json()
        elif r.status_code == 556:
            print(f"  ⚠ Cloudflare block (556) on {label or url}")
            return None
        elif r.status_code == 429:
            if _retry_count >= MAX_429_RETRIES:
                print(f"  ✗ Rate limit exceeded after {MAX_429_RETRIES} retries on {label or url}")
                return None
            wait_time = min(10 + _retry_count * 5, 30)
            print(f"  ⏳ Rate limited, waiting {wait_time}s... (retry {_retry_count + 1}/{MAX_429_RETRIES})")
            time.sleep(wait_time)
            return api_get(url, params, label, _retry_count=_retry_count + 1)
        elif r.status_code == 500:
            # Some Trendyol endpoints return 500 for unsupported transaction types
            # Log quietly and continue
            return None
        else:
            print(f"  ✗ {label}: HTTP {r.status_code} - {r.text[:200]}")
            return None
    except Exception as e:
        print(f"  ✗ {label}: {e}")
        return None


def epoch_ms(dt):
    """Convert datetime to epoch milliseconds."""
    return int(dt.timestamp() * 1000)


# ─── Data Fetchers ──────────────────────────────────────────────────
def fetch_products():
    """Fetch all products from Trendyol."""
    print("📦 Fetching products...")

    all_products = []

    # Try integration path first (international sellers)
    for page in range(200):
        data = api_get(
            f"{INTEGRATION_BASE}/product/sellers/{SELLER_ID}/products",
            params={"page": page, "size": 200},
            label=f"products-int p{page}",
        )
        if data and data.get("content"):
            all_products.extend(data["content"])
            if page >= data.get("totalPages", 1) - 1:
                break
        elif data and data.get("totalElements", 0) == 0 and page == 0:
            # Try sapigw path (domestic sellers)
            data2 = api_get(
                f"{SAPIGW_BASE}/suppliers/{SELLER_ID}/products",
                params={"page": 0, "size": 200, "approved": "true"},
                label="products-sapigw",
            )
            if data2 and data2.get("content"):
                all_products.extend(data2["content"])
                total_pages = data2.get("totalPages", 1)
                for p2 in range(1, total_pages):
                    d = api_get(
                        f"{SAPIGW_BASE}/suppliers/{SELLER_ID}/products",
                        params={"page": p2, "size": 200, "approved": "true"},
                        label=f"products-sapigw p{p2}",
                    )
                    if d and d.get("content"):
                        all_products.extend(d["content"])
            break
        else:
            break

    print(f"  → {len(all_products)} products found")
    return all_products


def fetch_settlements(start_date, end_date, transaction_types=None):
    """
    Fetch settlement records from Trendyol Finance API.
    Automatically splits into 15-day windows.

    IMPORTANT for international sellers:
      - storeFrontCode header is required (set globally in HEADERS)
      - transactionType must be sent ONE AT A TIME (not comma-separated)
      - size must be exactly 500 (smaller values return 400 error)
    """
    if transaction_types is None:
        transaction_types = [
            "Sale",
            "Return",
            "Discount",
            "DiscountCancel",
            "Coupon",
            "CouponCancel",
            "CommissionNegative",
            "CommissionPositive",
            "SellerRevenuePositive",
            "SellerRevenueNegative",
            "DeliveryFee",
        ]

    print(f"💰 Fetching settlements ({start_date.date()} → {end_date.date()})...")

    all_records = []
    window_start = start_date

    while window_start < end_date:
        window_end = min(window_start + timedelta(days=MAX_FINANCE_WINDOW_DAYS), end_date)

        # Trendyol API requires transactionType to be sent ONE AT A TIME
        for tx_type in transaction_types:
            for page in range(500):
                data = api_get(
                    f"{INTEGRATION_BASE}/finance/che/sellers/{SELLER_ID}/settlements",
                    params={
                        "startDate": epoch_ms(window_start),
                        "endDate": epoch_ms(window_end),
                        "transactionType": tx_type,
                        "page": page,
                        "size": FINANCE_PAGE_SIZE,
                    },
                    label=f"settlements/{tx_type} {window_start.date()}..{window_end.date()} p{page}",
                )

                if data and data.get("content"):
                    all_records.extend(data["content"])
                    if page >= data.get("totalPages", 1) - 1:
                        break
                else:
                    break

        window_start = window_end

    print(f"  → {len(all_records)} settlement records")
    return all_records


def fetch_other_financials(start_date, end_date):
    """Fetch other financial records (payments, invoices, shipping costs, etc.).

    Valid transactionTypes (as of 2026-04):
        CashAdvance, WireTransfer, IncomingTransfer, ReturnInvoice,
        CommissionAgreementInvoice, PaymentOrder, DeductionInvoices,
        FinancialItem, Stoppage, CreditNote, CommissionInvoice
    Note: 'ShippingInvoice' is NO LONGER valid. Shipping invoices are
          returned under 'DeductionInvoices' with transactionType='SHIPPING INVOICE-RO'.
    """
    print(f"📊 Fetching other financials ({start_date.date()} → {end_date.date()})...")

    # Only types that actually work and return useful data
    types = [
        "DeductionInvoices",       # SHIPPING INVOICE-RO + product fees (WRONG/MISSING/etc.)
        "ReturnInvoice",
        "CommissionInvoice",
        "CreditNote",
        "FinancialItem",
    ]

    all_records = []

    # Fetch for both RO and BG storefronts to capture Bulgaria shipping costs
    for storefront in ["RO", "BG"]:
        window_start = start_date
        while window_start < end_date:
            window_end = min(window_start + timedelta(days=MAX_FINANCE_WINDOW_DAYS), end_date)

            for tx_type in types:
                custom_headers = dict(session.headers)
                custom_headers["storeFrontCode"] = storefront
                try:
                    time.sleep(RATE_LIMIT_DELAY)
                    r = session.get(
                        f"{INTEGRATION_BASE}/finance/che/sellers/{SELLER_ID}/otherfinancials",
                        params={
                            "startDate": epoch_ms(window_start),
                            "endDate": epoch_ms(window_end),
                            "transactionType": tx_type,
                            "page": 0,
                            "size": FINANCE_PAGE_SIZE,
                        },
                        headers=custom_headers,
                        timeout=30,
                    )
                    if r.status_code == 200:
                        data = r.json()
                        if data.get("content"):
                            # Tag each record with the storefront it came from
                            for rec in data["content"]:
                                rec["_storefront"] = storefront
                            all_records.extend(data["content"])
                    # Silently skip 400/500 errors (invalid type for this storefront)
                except Exception:
                    pass

            window_start = window_end

        if storefront == "RO":
            print(f"    → {len(all_records)} records from RO storefront")
        else:
            ro_count = sum(1 for r in all_records if r.get("_storefront") == "RO")
            bg_count = len(all_records) - ro_count
            print(f"    → {bg_count} additional records from BG storefront")

    print(f"  → {len(all_records)} total financial records")
    return all_records


def fetch_orders(start_date, end_date):
    """
    Fetch shipment packages / orders via integration API.

    IMPORTANT:
      - Orders API returns incorrect results for date ranges > 7 days.
        We split into 7-day windows to get accurate data.
      - API filters on lastModifiedDate, not orderDate.
      - Each API entry is a shipment package (colet), NOT an order.
        One order can have multiple packages.
      - Deduplicate on shipmentPackageId across windows.
    """
    print(f"📋 Fetching orders ({start_date.date()} → {end_date.date()})...")

    all_orders = []
    seen_ids = set()  # Deduplicate across windows by packageId
    window_start = start_date

    while window_start < end_date:
        window_end = min(window_start + timedelta(days=MAX_ORDER_WINDOW_DAYS), end_date)

        for page in range(500):
            # Use integration path (sapigw returns 556 for international sellers)
            data = api_get(
                f"{INTEGRATION_BASE}/order/sellers/{SELLER_ID}/orders",
                params={
                    "startDate": epoch_ms(window_start),
                    "endDate": epoch_ms(window_end),
                    "page": page,
                    "size": 200,
                    "orderByField": "PackageLastModifiedDate",
                    "orderByDirection": "DESC",
                },
                label=f"orders {window_start.date()}..{window_end.date()} p{page}",
            )

            if data and data.get("content"):
                for order in data["content"]:
                    # Deduplicate on shipmentPackageId (each entry = 1 package/colet)
                    oid = order.get("shipmentPackageId") or order.get("orderNumber") or id(order)
                    if oid not in seen_ids:
                        seen_ids.add(oid)
                        all_orders.append(order)
                if page >= data.get("totalPages", 1) - 1:
                    break
            else:
                break

        window_start = window_end

    print(f"  → {len(all_orders)} packages (colete) found")
    return all_orders


# ─── Cost Loading ───────────────────────────────────────────────────
def load_costs(cost_file):
    """
    Load product costs from a CSV file.
    Expected columns: barcode (or sku), cost
    """
    costs = {}
    if not cost_file or not os.path.exists(cost_file):
        return costs

    print(f"📁 Loading costs from {cost_file}...")
    with open(cost_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = row.get("barcode") or row.get("sku") or row.get("Barcode") or row.get("SKU", "")
            cost = row.get("cost") or row.get("Cost") or row.get("pret_cost") or "0"
            try:
                costs[key.strip()] = float(cost.replace(",", "."))
            except (ValueError, AttributeError):
                pass

    print(f"  → {len(costs)} product costs loaded")
    return costs


def load_mapping_costs():
    """
    Load COGS from saved Trendyol→Shopify mapping (trendyol_mapping.json).
    Returns dict: {barcode: cost, sku: cost}
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Check data/ subdir first (where the API saves it), then script dir
    mapping_file = os.path.join(script_dir, "data", "trendyol_mapping.json")
    if not os.path.exists(mapping_file):
        mapping_file = os.path.join(script_dir, "trendyol_mapping.json")
    costs = {}
    if not os.path.exists(mapping_file):
        return costs

    print(f"📎 Loading COGS from mapping file...")
    try:
        with open(mapping_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data.get("mapping", []):
            cost = item.get("shopify_cost")
            if cost is None:
                continue
            try:
                cost_val = float(cost)
            except (ValueError, TypeError):
                continue
            bc = (item.get("trendyol_barcode") or "").strip()
            sku = (item.get("trendyol_sku") or "").strip()
            if bc:
                costs[bc] = cost_val
            if sku:
                costs[sku] = cost_val
        print(f"  → {len(costs)} COGS from Shopify mapping")
    except Exception as e:
        print(f"  ⚠ Error loading mapping: {e}")
    return costs


## fetch_shopify_costs() REMOVED — mapping JSON is now the primary COGS source.
## Shopify costs are fetched via the UI auto-mapping flow and stored in trendyol_mapping.json.
## This avoids a ~30-60s Shopify API call on every profitability calculation.


# ─── Profitability Calculation ──────────────────────────────────────
def calculate_profitability(settlements, products, costs, other_financials=None,
                            tva_rate=None, brand_map=None, transport_total=0):
    """
    Calculate profitability from settlement data.

    Formula (matching manual Excel calculation):
      Încasări = sum of credit (Valoarea facturii) for Sale — Return debt
      Profit brut = Încasări - COGS - Transport
      TVA de plată = (Încasări - COGS) * TVA_RATE
      Profit net = Profit brut - TVA de plată
    """
    if tva_rate is None:
        tva_rate = TVA_RATE
    if brand_map is None:
        brand_map = {}

    print(f"\n📈 Calculating profitability (TVA {tva_rate*100:.0f}%)...\n")

    # Index products by barcode
    product_map = {}
    for p in products:
        barcode = p.get("barcode", "")
        product_map[barcode] = {
            "title": p.get("title", "Unknown"),
            "barcode": barcode,
            "brandName": p.get("brandName", ""),
            "salePrice": p.get("salePrice", 0),
            "listPrice": p.get("listPrice", 0),
        }

    # Aggregate settlements by barcode
    barcode_data = {}
    total_commission = 0
    total_returns = 0
    total_discounts = 0

    for s in settlements:
        barcode = s.get("barcode", "unknown")
        tx_type = s.get("transactionType", "")
        credit = float(s.get("credit", 0))
        debt = float(s.get("debt", 0))
        commission_amount = float(s.get("commissionAmount", 0) or 0)
        commission_rate = float(s.get("commissionRate", 0) or 0)
        seller_revenue = float(s.get("sellerRevenue", 0) or 0)

        # Valoarea facturii = sellerRevenue + commissionAmount (confirmed per-item match)
        valoare_factura = seller_revenue + commission_amount

        if barcode not in barcode_data:
            pm = product_map.get(barcode, {})
            brand = (pm.get("brandName", "")
                     or brand_map.get(barcode, "")
                     or "Necunoscut")
            barcode_data[barcode] = {
                "barcode": barcode,
                "title": pm.get("title", barcode),
                "brand": brand,
                "sales_count": 0,
                "returns_count": 0,
                "valoare_factura_total": 0,  # Valoarea facturii (sr + cm)
                "return_amount": 0,
                "commission_total": 0,
                "discount_total": 0,
                "commission_rate": commission_rate,
                "cost_per_unit": costs.get(barcode, 0),
            }

        d = barcode_data[barcode]

        if tx_type in ("Satış", "Sale"):
            d["sales_count"] += 1
            d["valoare_factura_total"] += valoare_factura
            d["commission_total"] += commission_amount
            total_commission += commission_amount

        elif tx_type in ("İade", "Return"):
            d["returns_count"] += 1
            # Return amount = the original Valoarea facturii of the returned item
            return_val = seller_revenue + commission_amount
            d["return_amount"] += return_val
            total_returns += return_val

        elif tx_type in ("İndirim", "Discount"):
            d["discount_total"] += debt
            total_discounts += debt

    # ─── Per-product calculation ───
    results = []

    for barcode, d in barcode_data.items():
        # Încasări = Valoarea facturii NET de returnări (returnarea reversează vânzarea — vezi docstring "net profit")
        incasari = d["valoare_factura_total"] - d["return_amount"]

        # COGS CU TVA pe unități NETE (vândute − returnate; produsul returnat se restochează → costul revine)
        net_units = max(0, d["sales_count"] - d["returns_count"])
        cogs_cu_tva = d["cost_per_unit"] * net_units

        # fara TVA values
        incasari_fara_tva = incasari / (1 + tva_rate) if tva_rate > 0 else incasari
        cogs_fara_tva = cogs_cu_tva / (1 + tva_rate) if tva_rate > 0 else cogs_cu_tva

        # Commission on fara TVA (matches user's formula)
        commission_on_fara_tva = incasari_fara_tva * d["commission_rate"] / 100 if d["commission_rate"] > 0 else d["commission_total"]

        # Per-product profit = Valoarea fara TVA - COGS fara TVA - Commission
        profit_per_product = incasari_fara_tva - cogs_fara_tva - commission_on_fara_tva

        # Marjă pe încasări
        margin = (profit_per_product / incasari * 100) if incasari > 0 else 0

        results.append(
            {
                "Barcode": barcode,
                "Produs": d["title"],
                "Marcă": d["brand"],
                "Vânzări": d["sales_count"],
                "Returnări": d["returns_count"],
                "Încasări": round(incasari, 2),
                "Încasări fara TVA": round(incasari_fara_tva, 2),
                "COGS (cu TVA)": round(cogs_cu_tva, 2),
                "COGS fara TVA": round(cogs_fara_tva, 2),
                "Comision Trendyol": round(commission_on_fara_tva, 2),
                "Rată Comision (%)": round(d["commission_rate"], 2),
                "Transport (TVA 0%)": 0,  # Transport is total, allocated in summary
                "Profit Net": round(profit_per_product, 2),
                "Marjă (%)": round(margin, 1),
            }
        )

    # Sort by profit descending
    results.sort(key=lambda x: x["Profit Net"], reverse=True)

    # ─── Brand breakdown ───
    brand_data = {}
    for r in results:
        brand = r["Marcă"]
        if brand not in brand_data:
            brand_data[brand] = {
                "Marcă": brand,
                "Încasări": 0,
                "COGS (cu TVA)": 0,
                "Comisioane": 0,
                "Transport": 0,
                "Profit Brut": 0,
                "TVA": 0,
                "Profit Net": 0,
                "Vânzări": 0,
            }
        bd = brand_data[brand]
        bd["Încasări"] += r["Încasări"]
        bd["COGS (cu TVA)"] += r["COGS (cu TVA)"]
        bd["Comisioane"] += r["Comision Trendyol"]
        bd["Profit Net"] += r["Profit Net"]
        bd["Vânzări"] += r["Vânzări"]

    # Recalculate brand-level aggregates
    for bd in brand_data.values():
        bd["Profit Brut"] = bd["Încasări"] - bd["COGS (cu TVA)"] - bd["Comisioane"]
        bd["TVA"] = (bd["Încasări"] - bd["COGS (cu TVA)"]) * tva_rate
        bd["Profit Net"] = bd["Profit Brut"] - bd["TVA"]
        for k in ["Încasări", "COGS (cu TVA)", "Comisioane", "Transport", "Profit Brut", "TVA", "Profit Net"]:
            bd[k] = round(bd[k], 2)

    brand_results = sorted(brand_data.values(), key=lambda x: x["Profit Net"], reverse=True)

    # ─── Summary ───
    total_incasari = sum(r["Încasări"] for r in results)
    total_cogs = sum(r["COGS (cu TVA)"] for r in results)
    total_comisioane = sum(r["Comision Trendyol"] for r in results)

    # Formula: Profit brut = Încasări - COGS - Comisioane - Transport
    profit_brut = total_incasari - total_cogs - total_comisioane - transport_total
    # TVA = (Încasări - COGS) * TVA_RATE
    tva_de_plata = (total_incasari - total_cogs) * tva_rate
    # Profit net = Profit brut - TVA
    profit_net = profit_brut - tva_de_plata

    summary = {
        "Încasări": round(total_incasari, 2),
        "COGS (cu TVA)": round(total_cogs, 2),
        "Transport (TVA 0%)": round(transport_total, 2),
        "Profit Brut": round(profit_brut, 2),
        "TVA de Plată": round(tva_de_plata, 2),
        "Profit Net": round(profit_net, 2),
        "Marjă (%)": round(
            (profit_net / total_incasari * 100) if total_incasari > 0 else 0, 1
        ),
        "Comisioane Trendyol": round(total_comisioane, 2),
        "Returnări Total": round(total_returns, 2),
        "Reduceri Total": round(total_discounts, 2),
        "Nr. Produse Vândute": sum(r["Vânzări"] for r in results),
        "Nr. Returnări": sum(r["Returnări"] for r in results),
    }

    return results, summary, brand_results


# ─── Output ─────────────────────────────────────────────────────────
def print_results(results, summary, brand_results=None, output_file=None):
    """Print profitability results to console and optionally to Excel."""

    # ── Console output ──
    print("=" * 80)
    print("  📊 RAPORT PROFITABILITATE TRENDYOL")
    print("=" * 80)

    print("\n── Sumar General ──")
    for key, val in summary.items():
        if "Profit" in key:
            emoji = "💰"
        elif "TVA" in key:
            emoji = "🏛️"
        else:
            emoji = "📊"
        print(f"  {emoji} {key}: {val}")

    # ── Brand breakdown ──
    if brand_results:
        print(f"\n── Profitabilitate per Marcă ──")
        print(f"{'Marcă':<25} {'Încasări':>12} {'COGS':>10} {'Profit Brut':>12} {'TVA':>10} {'Profit Net':>12}")
        print("-" * 85)
        for bd in brand_results:
            print(
                f"  {bd['Marcă']:<23} {bd['Încasări']:>12.2f} {bd['COGS (cu TVA)']:>10.2f} "
                f"{bd['Profit Brut']:>12.2f} {bd['TVA']:>10.2f} {bd['Profit Net']:>12.2f}"
            )
        total_brand_profit = sum(bd["Profit Net"] for bd in brand_results)
        print("-" * 85)
        print(f"  {'Total general':<23} {'':>12} {'':>10} {'':>12} {'':>10} {total_brand_profit:>12.2f}")

    if results:
        print(f"\n── Top Produse (din {len(results)} total) ──")
        print(
            f"{'Barcode':<20} {'Produs':<30} {'Vânz':>5} {'Încasări':>10} {'COGS':>10} {'Profit':>10} {'Marjă':>7}"
        )
        print("-" * 95)

        for r in results[:20]:
            name = r["Produs"][:28]
            print(
                f"{r['Barcode']:<20} {name:<30} {r['Vânzări']:>5} "
                f"{r['Încasări']:>10.2f} {r['COGS (cu TVA)']:>10.2f} "
                f"{r['Profit Net']:>10.2f} {r['Marjă (%)']:>6.1f}%"
            )

        # Losers
        losers = [r for r in results if r["Profit Net"] < 0]
        if losers:
            print(f"\n── ⚠ Produse Neprofitabile ({len(losers)}) ──")
            for r in losers[:10]:
                print(
                    f"  ✗ {r['Barcode']} — {r['Produs'][:30]}: "
                    f"Profit {r['Profit Net']:.2f} TRY ({r['Marjă (%)']:.1f}%)"
                )
    else:
        print("\n  ℹ Nu s-au găsit tranzacții în perioada selectată.")

    print("\n" + "=" * 80)

    # ── Excel export ──
    if output_file and results:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()

            # Sheet 1: Per product
            ws = wb.active
            ws.title = "Profitabilitate"
            headers = list(results[0].keys())
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2B5797", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_idx, r in enumerate(results, 2):
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=r[h])
                    if "%" in h:
                        cell.number_format = "0.0"
                    elif any(kw in h for kw in ["Încasări", "COGS", "Profit", "TVA", "Comision", "Transport", "Venit"]):
                        cell.number_format = "#,##0.00"
                    if h == "Profit Net" and r[h] < 0:
                        cell.font = Font(color="FF0000", bold=True)

            # Sheet 2: Per brand
            if brand_results:
                ws2 = wb.create_sheet("Per Marcă")
                brand_headers = list(brand_results[0].keys())
                for col, h in enumerate(brand_headers, 1):
                    cell = ws2.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4B8B3B", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row_idx, bd in enumerate(brand_results, 2):
                    for col_idx, h in enumerate(brand_headers, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx, value=bd[h])
                        if any(kw in h for kw in ["Încasări", "COGS", "Profit", "TVA", "Transport"]):
                            cell.number_format = "#,##0.00"
                        if h == "Profit Net" and bd[h] < 0:
                            cell.font = Font(color="FF0000", bold=True)

            # Sheet 3: Summary
            ws3 = wb.create_sheet("Sumar")
            for row_idx, (key, val) in enumerate(summary.items(), 1):
                ws3.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                ws3.cell(row=row_idx, column=2, value=val)

            # Auto-width
            for ws_item in wb.worksheets:
                for col in ws_item.columns:
                    max_len = max(len(str(c.value or "")) for c in col)
                    ws_item.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

            wb.save(output_file)
            print(f"\n✅ Raport Excel salvat: {output_file}")

        except ImportError:
            csv_file = output_file.replace(".xlsx", ".csv")
            with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)
            print(f"\n✅ Raport CSV salvat: {csv_file}")
            print("  (Instalează openpyxl pentru export Excel: pip install openpyxl)")

    return results, summary


# ─── Main ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Trendyol Profitability Calculator")
    parser.add_argument("--days", type=int, default=30, help="Number of days to analyze (default: 30)")
    parser.add_argument("--start-date", type=str, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", type=str, help="End date (YYYY-MM-DD)")
    parser.add_argument("--cost-file", type=str, help="CSV file with product costs (barcode,cost)")
    parser.add_argument("--output", type=str, help="Output Excel file path")
    parser.add_argument("--tva-rate", type=float, default=0.21, help="TVA rate (default: 0.21)")
    parser.add_argument("--transport", type=float, default=None, help="Manual transport override (RON)")
    args = parser.parse_args()

    # Determine date range
    if args.start_date and args.end_date:
        start_date = datetime.strptime(args.start_date, "%Y-%m-%d")
        end_date = datetime.strptime(args.end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        period_label = f"{args.start_date} → {args.end_date}"
    else:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=args.days)
        period_label = f"ultimele {args.days} zile"
    output_file = args.output or f"trendyol_profit_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

    print("=" * 60)
    print("  🏪 TRENDYOL PROFITABILITY CALCULATOR")
    print(f"  Perioadă: {period_label}")
    print(f"  ({start_date.date()} → {end_date.date()})")
    print(f"  Seller ID: {SELLER_ID}")
    print(f"  TVA: {args.tva_rate*100:.0f}%")
    print("=" * 60)
    print()

    # 1. Fetch data
    products = fetch_products()

    # Fetch settlements with a wider range than requested — settlements are indexed
    # by payment date, not order date. We fetch 30 days before and after the period
    # to capture settlements paid outside the window but with orderDate inside it.
    fetch_start = start_date - timedelta(days=30)
    fetch_end = min(end_date + timedelta(days=30), datetime.now() + timedelta(days=1))
    print(f"\n🔍 Fetching settlements ({fetch_start.date()} → {fetch_end.date()}) to filter by order date...")
    all_settlements = fetch_settlements(fetch_start, fetch_end)

    # Filter settlements by orderDate to match requested period
    start_ts = int(start_date.timestamp() * 1000)
    end_ts = int(end_date.timestamp() * 1000)
    settlements = []
    for s in all_settlements:
        od = s.get("orderDate")
        if od and start_ts <= int(od) <= end_ts:
            settlements.append(s)
    print(f"  → {len(settlements)} settlements with orderDate in [{start_date.date()} → {end_date.date()}] (from {len(all_settlements)} total)")

    # Fetch shipping invoices (wider range — they are issued retroactively ~3 weeks after)
    # Shipping invoices have transactionType="SHIPPING INVOICE-RO" in the response
    shipping_fetch_start = start_date - timedelta(days=14)
    shipping_fetch_end = min(end_date + timedelta(days=45), datetime.now() + timedelta(days=1))
    print(f"\n🚚 Fetching shipping invoices ({shipping_fetch_start.date()} → {shipping_fetch_end.date()})...")
    other_financials = fetch_other_financials(shipping_fetch_start, shipping_fetch_end)
    shipping_invoices = []
    product_fees = []
    for rec in (other_financials or []):
        tx = rec.get("transactionType", "")
        debt = float(rec.get("debt", 0))
        td = rec.get("transactionDate", "")
        date_str = "?"
        if td:
            invoice_date = datetime.fromtimestamp(int(td)/1000)
            date_str = invoice_date.strftime("%Y-%m-%d")
        storefront = rec.get("_storefront", "RO")

        if "SHIPPING" in tx.upper():
            shipping_invoices.append({"date": date_str, "amount": debt, "type": tx, "sf": storefront})
            print(f"    🚚 {date_str}: {debt:.2f} RON ({tx}) [{storefront}]")
        elif any(kw in tx.upper() for kw in ["WRONG", "MISSING", "DEFECTIVE", "UNSUPPLIED"]):
            product_fees.append({"date": date_str, "amount": debt, "type": tx, "sf": storefront})
            print(f"    ⚠️  {date_str}: {debt:.2f} RON ({tx}) [{storefront}]")

    # Use manual transport if provided, otherwise sum invoices within the period
    if args.transport is not None:
        transport_total = args.transport
        print(f"  → Transport (manual): {transport_total:.2f} RON")
    else:
        # Sum shipping invoices where invoice date falls within the selected period
        transport_total = 0
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")
        for inv in shipping_invoices:
            if inv["date"] >= start_str and inv["date"] <= end_str:
                transport_total += inv["amount"]
                print(f"    ✅ {inv['date']}: {inv['amount']:.2f} RON ({inv['type']})")
            else:
                print(f"    ⏭️  {inv['date']}: {inv['amount']:.2f} RON (outside period)")

        # Also sum product fees within period
        fees_total = 0
        for fee in product_fees:
            if fee["date"] >= start_str and fee["date"] <= end_str:
                fees_total += fee["amount"]

        print(f"  → Transport ({start_date.date()} - {end_date.date()}): {transport_total:.2f} RON")
        print(f"  → Penalizări produse: {fees_total:.2f} RON")
        if transport_total == 0:
            print(f"  ⚠️  Nicio factură de transport găsită. Poți seta manual cu --transport")
        # Add fees to transport total (they are deductions too)
        transport_total += fees_total

    # 2. Load costs — Mapping JSON is PRIMARY (populated via UI auto-mapping from Shopify)
    costs = load_mapping_costs()  # COGS from trendyol_mapping.json (Shopify-matched)
    csv_costs = load_costs(args.cost_file)  # Optional CSV override
    costs.update(csv_costs)  # CSV always overrides
    print(f"\n  📊 Total COGS entries: {len(costs)}")

    # 3. Load brand map from mapping file
    brand_map = {}
    script_dir = os.path.dirname(os.path.abspath(__file__))
    mapping_file = os.path.join(script_dir, "data", "trendyol_mapping.json")
    if not os.path.exists(mapping_file):
        mapping_file = os.path.join(script_dir, "trendyol_mapping.json")
    if os.path.exists(mapping_file):
        try:
            with open(mapping_file, "r", encoding="utf-8") as f:
                mdata = json.load(f)
            for item in mdata.get("mapping", []):
                bc = (item.get("trendyol_barcode") or "").strip()
                brand = (item.get("shopify_store") or "").strip()
                if bc and brand:
                    brand_map[bc] = brand
        except Exception:
            pass

    # 4. Calculate profitability
    results, summary, brand_results = calculate_profitability(
        settlements, products, costs, other_financials,
        tva_rate=args.tva_rate, brand_map=brand_map,
        transport_total=transport_total,
    )

    # 5. Output
    print_results(results, summary, brand_results, output_file)

    # 6. Save to history
    history_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "trendyol_results_history.json")
    history = []
    if os.path.exists(history_file):
        try:
            with open(history_file, "r", encoding="utf-8") as f:
                history = json.load(f)
        except Exception:
            history = []

    history_entry = {
        "timestamp": datetime.now().isoformat(),
        "period": {
            "start": start_date.strftime("%Y-%m-%d"),
            "end": end_date.strftime("%Y-%m-%d"),
            "label": period_label,
        },
        "summary": summary,
        "brands": brand_results,
        "products": results[:50],
    }
    history.append(history_entry)
    # Keep last 50 entries
    history = history[-50:]
    try:
        with open(history_file, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        print(f"\n📁 Istoric salvat ({len(history)} intrări)")
    except Exception as e:
        print(f"  ⚠ Eroare salvare istoric: {e}")

    # 7. Return JSON for dashboard integration
    output = {
        "summary": summary,
        "products": results[:50],
        "brands": brand_results,
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
            "label": period_label,
        },
    }
    print(f"\n###TRENDYOL_RESULTS###{json.dumps(output, ensure_ascii=False)}")

    return results, summary, brand_results


if __name__ == "__main__":
    main()

