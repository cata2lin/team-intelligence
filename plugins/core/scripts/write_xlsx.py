# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "psycopg2-binary>=2.9"]
# ///
"""write_xlsx — run a Postgres SELECT and write a styled .xlsx workbook.

Secrets come from the SharedClaude `secrets` table, not a .env file: this script
imports the kb_env shim and calls load_secrets_into_env() up front, which sets
os.environ (DATABASE_URL_*, ...) from $KB_DATABASE_URL.

The DB connection name (--db) is the *env key* holding a Postgres connection
string (e.g. DATABASE_URL_METRICS), so the actual credential never appears on the
command line. The query runs in a READ ONLY transaction.

Usage:
    uv run write_xlsx.py \
        --db DATABASE_URL_METRICS \
        --sql "SELECT id, name, created_at FROM brands ORDER BY id" \
        --out "$NAS_ROOT/exports/brands.xlsx"

Options:
    --db KEY        env var holding the Postgres connection string (required)
    --sql QUERY     the SELECT to run; mutually exclusive with --sql-file
    --sql-file PATH read the SQL from a file instead of --sql
    --out PATH      destination .xlsx (default: out.xlsx)
    --sheet NAME    worksheet title (default: data)

> Author: Arona core
"""
import argparse
import os
import sys

from kb_env import load_secrets_into_env

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)
MAX_COL_WIDTH = 60


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        prog="write_xlsx.py",
        description="Run a Postgres SELECT and write a styled .xlsx workbook.",
    )
    p.add_argument("--db", required=True,
                   help="env var name holding the Postgres connection string "
                        "(e.g. DATABASE_URL_METRICS)")
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--sql", help="the SELECT query to run")
    src.add_argument("--sql-file", help="read the SQL from this file instead")
    p.add_argument("--out", default="out.xlsx", help="destination .xlsx path")
    p.add_argument("--sheet", default="data", help="worksheet title")
    return p.parse_args(argv)


def resolve_dsn(db_key):
    """Map the --db env-key to an actual connection string."""
    dsn = os.environ.get(db_key)
    if not dsn:
        sys.stderr.write(
            f"[write_xlsx] no connection string in env var {db_key!r}. "
            f"Pass the key that holds a DATABASE_URL (is $KB_DATABASE_URL set?).\n"
        )
        sys.exit(2)
    return dsn


def fetch(dsn, sql):
    """Run sql read-only; return (headers, rows)."""
    conn = psycopg2.connect(dsn, connect_timeout=15)
    try:
        conn.set_session(readonly=True)
        with conn.cursor() as cur:
            cur.execute(sql)
            if cur.description is None:
                sys.stderr.write("[write_xlsx] query returned no result set.\n")
                sys.exit(2)
            headers = [d.name for d in cur.description]
            rows = cur.fetchall()
    finally:
        conn.close()
    return headers, rows


def build_workbook(headers, rows, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for r in rows:
        ws.append(list(r))

    # Styled header: bold, grey fill, frozen.
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"

    # Auto-size columns from their longest cell (capped).
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, MAX_COL_WIDTH)

    return wb


def main(argv=None):
    args = parse_args(argv)

    # Pull DATABASE_URL_* (and friends) from the SharedClaude secrets table.
    load_secrets_into_env()

    sql = args.sql
    if args.sql_file:
        with open(args.sql_file, "r", encoding="utf-8") as fh:
            sql = fh.read()
    if not sql or not sql.strip():
        sys.stderr.write("[write_xlsx] empty SQL.\n")
        sys.exit(2)

    dsn = resolve_dsn(args.db)
    headers, rows = fetch(dsn, sql)

    out_dir = os.path.dirname(os.path.abspath(args.out))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    wb = build_workbook(headers, rows, args.sheet)
    wb.save(args.out)

    print(f"wrote {args.out} ({len(rows)} rows, {len(headers)} cols)")


if __name__ == "__main__":
    main()
