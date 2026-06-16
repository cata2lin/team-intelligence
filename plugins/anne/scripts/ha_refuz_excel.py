# /// script
# requires-python = ">=3.10"
# dependencies = ["paramiko>=3.0", "psycopg2-binary>=2.9", "openpyxl>=3.1"]
# ///
"""
Exporta analiza refuz/retur HA in Excel, cu 3 sheet-uri:
  1. Rezumat  — agregat per perioada de timp
  2. Per produs — toate perioadele, culori, trend
  3. Schimbari — produse cu modificare semnificativa (>5pp)

Usage:
  uv run ha_refuz_excel.py [--out fisier.xlsx] [--min-orders N]
"""
import argparse, json, os, sys
import psycopg2, paramiko
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

SSH_HOST_KEY  = "PROFIT_SSH_HOST"
SSH_USER_KEY  = "PROFIT_SSH_USER"
SSH_PASS_KEY  = "PROFIT_SSH_PASS"
REMOTE_PYTHON = "/root/Scripturi/.venv/bin/python"

QUERY_SCRIPT = r"""
import sqlite3, json, os
os.chdir('/root/Scripturi')
db = sqlite3.connect('data/profitability.db')

# Refuz = orice colet intors (status_category='Refuzata'), indiferent de shopify_delivery_status.
# refuz_pct = Refuzate / (Livrate + Refuzate) * 100
# Numitor = colete care au plecat efectiv (Anulate nu au plecat).

PERIODS = [
    ("7d",  "date('now', '-7 days')",  None),
    ("30d", "date('now', '-30 days')", "date('now', '-7 days')"),
    ("90d", "date('now', '-90 days')", "date('now', '-30 days')"),
    ("old", None,                      "date('now', '-90 days')"),
]

results = {}
for period_name, date_from, date_to in PERIODS:
    where_parts = ["skus LIKE 'HA-%'",
                   "status_category IN ('Refuzata', 'Livrata', 'Anulata')"]
    if date_from:
        where_parts.append(f"DATE(created_at) >= {date_from}")
    if date_to:
        where_parts.append(f"DATE(created_at) < {date_to}")
    where_clause = " AND ".join(where_parts)

    rows = db.execute(f'''
        SELECT skus, status_category, COUNT(*) as cnt
        FROM profit_orders
        WHERE {where_clause}
        GROUP BY skus, status_category
    ''').fetchall()

    data = {}
    for skus_val, cat, cnt in rows:
        sku = skus_val.strip()
        if ";" in sku:
            continue
        if sku not in data:
            data[sku] = {"livrata": 0, "refuz": 0, "anulata": 0}
        if cat == 'Livrata':
            data[sku]["livrata"] += cnt
        elif cat == 'Anulata':
            data[sku]["anulata"] += cnt
        elif cat == 'Refuzata':
            data[sku]["refuz"] += cnt

    period_out = {}
    for sku, v in data.items():
        liv, ref, an = v["livrata"], v["refuz"], v["anulata"]
        trimise = liv + ref
        period_out[sku] = {
            "livrata": liv, "refuz": ref, "anulata": an,
            "total": trimise,
            "refuz_pct": round(ref / trimise * 100, 2) if trimise else None,
        }
    results[period_name] = period_out

print(json.dumps(results))
"""


def kb_get(key):
    url = os.environ.get("KB_DATABASE_URL")
    if not url:
        return None
    try:
        conn = psycopg2.connect(url, connect_timeout=10)
        with conn, conn.cursor() as cur:
            cur.execute("SELECT value FROM secrets WHERE key=%s", (key,))
            row = cur.fetchone()
            return row[0] if row else None
    except Exception:
        return None
    finally:
        conn.close()


def run_remote():
    host = kb_get(SSH_HOST_KEY) or "84.46.242.181"
    user = kb_get(SSH_USER_KEY) or "root"
    pwd  = kb_get(SSH_PASS_KEY)
    if not pwd:
        sys.exit(f"EROARE: secretul '{SSH_PASS_KEY}' lipseste din KB.")
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(host, username=user, password=pwd, timeout=30)
    sftp = client.open_sftp()
    with sftp.open("/tmp/_ha_trend_xl.py", "w") as f:
        f.write(QUERY_SCRIPT)
    sftp.close()
    _, stdout, stderr = client.exec_command(
        f"{REMOTE_PYTHON} /tmp/_ha_trend_xl.py", timeout=120)
    raw  = stdout.read().decode()
    errs = stderr.read().decode().strip()
    client.close()
    if errs:
        print(f"[remote stderr]: {errs[:300]}", file=sys.stderr)
    return json.loads(raw)


# ── Styles ────────────────────────────────────────────────────────────────────

PERIODS_ORDER  = ["old", "90d", "30d", "7d"]
PERIOD_LABELS  = {
    "old": "91+ zile in urma",
    "90d": "31-90 zile in urma",
    "30d": "8-30 zile in urma",
    "7d":  "Ultimele 7 zile",
}

HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
HDR2_FILL  = PatternFill("solid", fgColor="2E5FA3")   # mid blue
SUB_FILL   = PatternFill("solid", fgColor="D9E1F2")   # light blue
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL  = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
HDR2_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BOLD       = Font(bold=True, name="Calibri", size=10)
NORMAL     = Font(name="Calibri", size=10)
SMALL      = Font(name="Calibri", size=9, color="595959")

CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center")
RIGHT      = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def pct_str(val):
    if val is None:
        return None
    return val / 100.0   # store as float, format as % in Excel

def trend_label(old_val, new_val):
    if old_val is None or new_val is None:
        return ""
    diff = new_val - old_val
    if abs(diff) < 0.5:
        return "-> stabil"
    sign = "-" if diff < 0 else "+"
    return f"{sign}{abs(diff):.1f}pp"

def trend_fill(old_val, new_val):
    if old_val is None or new_val is None:
        return None
    diff = new_val - old_val
    if diff <= -5:
        return GREEN_FILL
    if diff >= 5:
        return RED_FILL
    return None


# ── Sheet 1: Rezumat agregat ──────────────────────────────────────────────────

def build_sheet_rezumat(wb, data):
    ws = wb.create_sheet("Rezumat agregat")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Evolutie Rata de Refuz (colete intorse) — Produse HA"
    c.font = Font(bold=True, size=14, name="Calibri", color="1F3864")
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = ("Refuz% = colete intorse (Refuzata) / colete trimise (Livrate + Refuzate)  |  "
               "Incluse: refuz la usa + retur fizic dupa primire — orice colet care nu a ramas la client")
    c.font = SMALL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    headers = ["Perioada", "Trimise (Liv+Ref)", "Livrate", "Intorse (Refuz)", "Anulate", "Refuz%"]
    col_widths = [24, 20, 12, 18, 12, 10]

    row = 4
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[row].height = 22

    # Aggregate each period
    agg_rows = []
    for p in PERIODS_ORDER:
        pd = data.get(p, {})
        liv = ref = an = 0
        for sku, v in pd.items():
            liv += v["livrata"]; ref += v["refuz"]; an += v["anulata"]
        trimise = liv + ref
        agg_rows.append({
            "label": PERIOD_LABELS[p],
            "trimise": trimise,
            "livrata": liv, "refuz": ref, "anulata": an,
            "refuz_pct": ref / trimise * 100 if trimise else None,
        })

    fills = [GREY_FILL, WHITE_FILL, GREY_FILL, WHITE_FILL]
    for i, a in enumerate(agg_rows):
        row += 1
        fill = fills[i]
        vals = [a["label"], a["trimise"], a["livrata"], a["refuz"], a["anulata"],
                pct_str(a["refuz_pct"])]
        fmts = [None, "#,##0", "#,##0", "#,##0", "#,##0", "0.0%"]
        for col_i, (val, fmt) in enumerate(zip(vals, fmts), start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.fill = fill
            c.font = NORMAL
            c.alignment = CENTER if col_i > 1 else LEFT
            c.border = thin_border()
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[row].height = 18

    # Trend row — compare 90d vs 30d (most data)
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    old_r = agg_rows[1]["refuz_pct"]  # 90d
    new_r = agg_rows[2]["refuz_pct"]  # 30d
    if old_r and new_r:
        diff = new_r - old_r
        sign = "scazut" if diff < 0 else "crescut"
        msg = (f"Trend refuz (31-90 zile vs 8-30 zile): rata a {sign} cu "
               f"{abs(diff):.1f}pp  ({old_r:.1f}% -> {new_r:.1f}%)")
    else:
        msg = "Date insuficiente pentru trend."
    c = ws[f"A{row}"]
    c.value = msg
    c.font = Font(bold=True, size=11, name="Calibri",
                  color="375623" if (old_r and new_r and new_r < old_r) else "843C0C")
    c.alignment = LEFT
    ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A5"
    return ws


# ── Sheet 2: Per produs ───────────────────────────────────────────────────────

def build_sheet_per_produs(wb, data, min_orders, sort_metric="retur"):
    ws = wb.create_sheet("Per produs")
    ws.sheet_view.showGridLines = False

    # Collect all SKUs with enough total orders
    all_skus = set()
    for pd in data.values():
        all_skus.update(pd.keys())

    def total_orders(sku):
        return sum(data.get(p, {}).get(sku, {}).get("total", 0) for p in PERIODS_ORDER)

    sort_key = f"{sort_metric}_pct"
    skus = sorted([s for s in all_skus if total_orders(s) >= min_orders],
                  key=lambda s: (data.get("90d", {}).get(s, {}).get(sort_key) or
                                 data.get("30d", {}).get(s, {}).get(sort_key) or 0),
                  reverse=True)

    # Build header rows
    # Row 1: title
    total_cols = 2 + len(PERIODS_ORDER) * 5 + 2  # SKU + total + (Liv,Ref,Ret,An,R%,Rtr%) * 4 periods + trend refuz + trend retur
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = f"Evolutie per produs HA — {len(skus)} SKU-uri cu minim {min_orders} comenzi totale"
    c.font = Font(bold=True, size=13, name="Calibri", color="1F3864")
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Row 2: period group headers
    row = 2
    ws.cell(row=row, column=1, value="SKU").fill = HDR_FILL
    ws.cell(row=row, column=1).font = HDR_FONT
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = thin_border()
    ws.column_dimensions["A"].width = 13

    ws.cell(row=row, column=2, value="Total comenzi").fill = HDR_FILL
    ws.cell(row=row, column=2).font = HDR_FONT
    ws.cell(row=row, column=2).alignment = CENTER
    ws.cell(row=row, column=2).border = thin_border()
    ws.column_dimensions["B"].width = 14

    col = 3
    COLS_PER_PERIOD = 4  # Livrate, Intorse, Anulate, Refuz%
    for p in PERIODS_ORDER:
        end_col = col + COLS_PER_PERIOD - 1
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=col, value=PERIOD_LABELS[p])
        c.fill = HDR2_FILL
        c.font = HDR2_FONT
        c.alignment = CENTER
        c.border = thin_border()
        col += COLS_PER_PERIOD

    # Trend column (refuz%)
    c = ws.cell(row=row, column=col, value="Trend Refuz%")
    c.fill = PatternFill("solid", fgColor="4A235A")
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = CENTER
    c.border = thin_border()
    ws.column_dimensions[get_column_letter(col)].width = 14
    col += 1

    ws.row_dimensions[row].height = 32

    # Row 3: sub-headers
    row = 3
    ws.cell(row=row, column=1).fill = SUB_FILL
    ws.cell(row=row, column=2, value="(toate perioadele)").fill = SUB_FILL
    ws.cell(row=row, column=2).font = SMALL
    ws.cell(row=row, column=2).alignment = CENTER

    col = 3
    sub_headers = ["Livrate", "Intorse", "Anulate", "Refuz%"]
    sub_widths  = [9, 9, 9, 9]
    for p in PERIODS_ORDER:
        for h, w in zip(sub_headers, sub_widths):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = SUB_FILL
            c.font = Font(bold=True, name="Calibri", size=9, color="1F3864")
            c.alignment = CENTER
            c.border = thin_border()
            ws.column_dimensions[get_column_letter(col)].width = w
            col += 1

    ws.cell(row=row, column=col-2)  # already set trend cols above
    ws.row_dimensions[row].height = 18

    # Data rows
    row_fills = [WHITE_FILL, GREY_FILL]
    for sku_i, sku in enumerate(skus):
        row = 4 + sku_i
        fill = row_fills[sku_i % 2]

        # SKU
        c = ws.cell(row=row, column=1, value=sku)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = fill
        c.alignment = LEFT
        c.border = thin_border()

        # Total all periods
        tot = total_orders(sku)
        c = ws.cell(row=row, column=2, value=tot)
        c.fill = fill; c.font = NORMAL; c.alignment = CENTER
        c.border = thin_border(); c.number_format = "#,##0"

        col = 3
        refuz_pcts = []
        for p in PERIODS_ORDER:
            v = data.get(p, {}).get(sku)
            if v:
                vals = [v["livrata"], v["refuz"], v["anulata"], pct_str(v["refuz_pct"])]
                fmts = ["#,##0", "#,##0", "#,##0", "0.0%"]
                refuz_pcts.append(v["refuz_pct"])
            else:
                vals = [None, None, None, None]
                fmts = [None] * 4
                refuz_pcts.append(None)

            for val, fmt in zip(vals, fmts):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill; c.font = NORMAL; c.alignment = CENTER
                c.border = thin_border()
                if fmt and val is not None:
                    c.number_format = fmt
                col += 1

        # Single trend column (refuz%)
        non_null_r = [(i, v) for i, v in enumerate(refuz_pcts) if v is not None]
        if len(non_null_r) >= 2:
            old_v = non_null_r[0][1]
            new_v = non_null_r[-1][1]
            label = trend_label(old_v, new_v)
            tfill = trend_fill(old_v, new_v) or fill
        else:
            label = "—"
            tfill = fill

        c = ws.cell(row=row, column=col, value=label)
        c.fill = tfill
        c.font = Font(bold=(tfill != fill), name="Calibri", size=10)
        c.alignment = CENTER
        c.border = thin_border()

        ws.row_dimensions[row].height = 16

    # Conditional formatting on Refuz% columns (4th sub-col = col+3)
    refuz_pct_cols = []
    col = 3
    for _ in PERIODS_ORDER:
        refuz_pct_cols.append(get_column_letter(col + 3))  # 4th sub-col = Refuz%
        col += COLS_PER_PERIOD

    last_data_row = 3 + len(skus)
    for letter in refuz_pct_cols:
        rng = f"{letter}4:{letter}{last_data_row}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0,   start_color="63BE7B",
            mid_type="num",   mid_value=0.25,  mid_color="FFEB84",
            end_type="num",   end_value=0.50,  end_color="F8696B",
        ))

    if False:
        ws.conditional_formatting.add("", ColorScaleRule(
            start_type="num", start_value=0,    start_color="63BE7B",
            mid_type="num",   mid_value=0.05,   mid_color="FFEB84",
            end_type="num",   end_value=0.15,   end_color="F8696B",
        ))

    ws.freeze_panes = "B4"
    return ws


# ── Sheet 3: Schimbari semnificative ─────────────────────────────────────────

def build_sheet_schimbari(wb, data, min_orders, metric="refuz"):
    ws = wb.create_sheet("Schimbari semnificative")
    ws.sheet_view.showGridLines = False

    baseline_period, recent_period = "90d", "30d"
    baseline_label, recent_label   = "31-90 zile in urma", "8-30 zile in urma"
    comparison_note = ""

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = (f"Produse HA cu schimbari semnificative in rata de {metric} "
               f"({baseline_label} vs {recent_label})")
    c.font = Font(bold=True, size=13, name="Calibri", color="1F3864")
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    if comparison_note:
        ws.merge_cells("A2:H2")
        c = ws["A2"]
        c.value = comparison_note
        c.font = Font(bold=True, size=10, name="Calibri", color="843C0C")
        c.fill = PatternFill("solid", fgColor="FCE4D6")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 40

    headers = ["SKU", f"Baseline ({baseline_label})", f"Recent ({recent_label})",
               "Diferenta (pp)", "Trend", "Total comenzi (baseline)",
               "Total comenzi (recent)", "Interpretare"]
    col_widths = [13, 22, 22, 16, 12, 24, 22, 40]

    row = 3
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=row, column=col_i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[row].height = 22

    all_skus = set()
    for pd in data.values():
        all_skus.update(pd.keys())

    def total_orders(sku):
        return sum(data.get(p, {}).get(sku, {}).get("total", 0) for p in PERIODS_ORDER)

    skus_enough = [s for s in all_skus if total_orders(s) >= min_orders]

    metric_key = f"{metric}_pct"
    changes = []
    for sku in skus_enough:
        old_v = data.get(baseline_period, {}).get(sku, {}).get(metric_key)
        new_v = data.get(recent_period,   {}).get(sku, {}).get(metric_key)
        if old_v is None or new_v is None:
            continue
        diff = new_v - old_v
        if abs(diff) < 5:
            continue
        old_n = data.get(baseline_period, {}).get(sku, {}).get("total", 0)
        new_n = data.get(recent_period,   {}).get(sku, {}).get("total", 0)
        changes.append((sku, old_v, new_v, diff, old_n, new_n))

    improved = sorted([c for c in changes if c[3] < 0], key=lambda x: x[3])
    worsened = sorted([c for c in changes if c[3] > 0], key=lambda x: x[3], reverse=True)

    def interpret(sku, old_v, new_v, diff, old_n, new_n):
        if new_n < 10:
            return "Date insuficiente in perioada recenta — interpret cu precautie"
        if diff <= -10:
            return "Imbunatatire semnificativa — posibil schimbare pozitiva in produs/descriere"
        if diff <= -5:
            return "Imbunatatire moderata — monitorizare recomandata"
        if diff >= 20:
            return "Inrautatire severa — actiune urgenta necesara"
        if diff >= 10:
            return "Inrautatire semnificativa — investigare recomandata"
        return "Inrautatire moderata — monitorizare recomandata"

    def write_group_header(title, fill_color):
        nonlocal row
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.font = Font(bold=True, size=11, name="Calibri",
                      color="FFFFFF" if fill_color != "E2EFDA" else "375623")
        c.alignment = LEFT
        ws.row_dimensions[row].height = 20
        row += 1

    def write_entry(entry, fill):
        nonlocal row
        sku, old_v, new_v, diff, old_n, new_n = entry
        interp = interpret(sku, old_v, new_v, diff, old_n, new_n)
        vals = [sku, f"{old_v:.1f}%", f"{new_v:.1f}%",
                f"{diff:+.1f}pp", "vv SCADE" if diff < 0 else "^^ CRESTE",
                old_n, new_n, interp]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.fill = fill; c.font = NORMAL; c.alignment = CENTER
            c.border = thin_border()
            if col_i == 1:
                c.font = Font(bold=True, name="Calibri", size=10)
                c.alignment = LEFT
            if col_i == 8:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

    write_group_header(f"IMBUNATATITE — rata {metric} a SCAZUT cu >=5pp  ({len(improved)} produse)", "375623")
    for i, entry in enumerate(improved):
        write_entry(entry, GREEN_FILL if i % 2 == 0 else PatternFill("solid", fgColor="C6EFCE"))

    if not improved:
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value="Niciun produs cu imbunatatire semnificativa").font = SMALL
        row += 1

    write_group_header(f"INRAUTATITE — rata {metric} a CRESCUT cu >=5pp  ({len(worsened)} produse)", "843C0C")
    for i, entry in enumerate(worsened):
        write_entry(entry, RED_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FADADD"))

    if not worsened:
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value="Niciun produs cu inrautatire semnificativa").font = SMALL

    ws.freeze_panes = "A4"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out",        default="ha_refuz_retur_trend.xlsx",
                    help="Fisier Excel de output (default: ha_refuz_retur_trend.xlsx)")
    ap.add_argument("--min-orders", type=int, default=30,
                    help="Minim comenzi totale pentru a include un produs (default 30)")
    args = ap.parse_args()

    print("Se conecteaza si ruleaza interogarea pe 4 ferestre de timp...", file=sys.stderr)
    data = run_remote()

    print("Se construieste Excel-ul...", file=sys.stderr)
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet_rezumat(wb, data)
    build_sheet_per_produs(wb, data, args.min_orders, sort_metric="retur")
    build_sheet_schimbari(wb, data, args.min_orders, metric="refuz")

    out_path = args.out
    wb.save(out_path)
    print(f"\nFisier salvat: {os.path.abspath(out_path)}", file=sys.stderr)
    print(os.path.abspath(out_path))


if __name__ == "__main__":
    main()
