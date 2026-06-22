"""
Exporter XLSX compatibil cu importul SmartBill e-Transport.
Generează un fișier cu coloanele exacte cerute de SmartBill.
"""
import os
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from etransport.models.shipment import Shipment
from etransport import config


def export_smartbill_xlsx(
    shipment: Shipment,
    output_path: str,
    sheet_name: Optional[str] = None,
) -> str:
    """
    Generează fișierul XLSX pentru import SmartBill e-Transport.
    
    Coloanele exportate (în ordine):
    1. Denumire produs
    2. Greutate netă
    3. Greutate brută
    4. Cod produs
    5. U.M. produs
    6. Moneda
    7. Cantitate
    8. Cod standard pentru U.M.
    9. Pret unitar fara TVA
    10. Cod tarifar (N.C.)
    11. Scop operatiune
    
    Args:
        shipment: Modelul de expediere complet
        output_path: Calea fișierului de output
        sheet_name: Numele sheet-ului (opțional)
        
    Returns:
        Calea absolută a fișierului generat
    """
    sheet_name = sheet_name or config.XLSX_SHEET_NAME
    
    # Asigurăm directorul de output
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    
    # Construim datele rând cu rând, dar EXCLUSIV pentru sursa INVOICE
    rows = []
    for product in shipment.products:
        if product.source_type != "invoice":
            continue
        row = _build_row(product, shipment)
        rows.append(row)
    
    # Creăm workbook-ul
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header
    columns = config.SMARTBILL_XLSX_COLUMNS
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Date
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Formatare specifică per coloană
            if col_name in ("Greutate netă", "Greutate brută"):
                cell.number_format = '0.000'
            elif col_name == "Pret unitar fara TVA":
                cell.number_format = '0.00'
            elif col_name == "Cantitate":
                cell.number_format = '0'
            elif col_name == "Cod tarifar (N.C.)":
                # IMPORTANT: NC8 trebuie păstrat ca text pentru a nu pierde zerouri
                cell.number_format = '@'
            elif col_name == "Cod produs":
                cell.number_format = '@'
            
            cell.alignment = Alignment(vertical="center")
    
    # Auto-width coloane
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)
    
    # Salvăm
    wb.save(output_path)
    return os.path.abspath(output_path)


def _build_row(product, shipment: Shipment) -> dict:
    """Construiește un rând de date conform coloanelor SmartBill."""
    
    # Scopul operațiunii — verificăm override-uri
    purpose = product.operation_purpose_code
    if shipment.operation_type in config.OPERATION_PURPOSE_OVERRIDES:
        override = config.OPERATION_PURPOSE_OVERRIDES[shipment.operation_type]
        purpose = override["code"]
        
    # Validare si curatare de cod NC / Tarifar
    nc8 = product.nc_code_8
    if getattr(product, "tariff_code_db_match", None):
        product.nc_code_override_applied = True
        nc8 = product.tariff_code_db_match
        product.nc_code_status = f"db_{product.tariff_code_match_method}"
    elif nc8 in config.NC_CODE_OVERRIDES:
        product.nc_code_override_applied = True
        nc8 = config.NC_CODE_OVERRIDES[nc8]
        product.nc_code_status = "override"
    elif not nc8 or len(nc8) != 8 or not nc8.isdigit():
        product.nc_code_status = "problematic"
    else:
        product.nc_code_status = "valid"

    # Calculare greutati unitare cu protectie ZeroDivisionError
    q = product.quantity if product.quantity > 0 else 1.0
    net_unit = product.net_weight_kg / q
    gross_unit = product.gross_weight_kg / q
    
    # Valuta si pret
    if config.EXPORT_CURRENCY_MODE == "original":
        currency_written = product.currency
        price_written = product.unit_price_original
    else:
        currency_written = config.OUTPUT_CURRENCY
        price_written = product.unit_price_ron
        
    # Populare modele de audit pe produs (referința obiect va fi vazuta de audit_json)
    product.export_weight_net_unit = net_unit
    product.export_weight_gross_unit = gross_unit
    product.export_unit_price_written = price_written
    product.export_currency_written = currency_written
    product.export_nc_code_written = nc8
    
    return {
        "Denumire produs": product.product_name_export,
        "Greutate netă": product.export_weight_net_unit,
        "Greutate brută": product.export_weight_gross_unit,
        "Cod produs": product.product_code or "",
        "U.M. produs": product.display_unit,
        "Moneda": product.export_currency_written,
        "Cantitate": int(product.quantity) if product.quantity == int(product.quantity) else product.quantity,
        "Cod standard pentru U.M.": product.standard_unit_code,
        "Pret unitar fara TVA": product.export_unit_price_written,
        "Cod tarifar (N.C.)": product.export_nc_code_written,  # 8 cifre, ca string
        "Scop operatiune": purpose,
    }
